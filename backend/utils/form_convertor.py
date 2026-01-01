"""
Excel Form Converter for GIU HR Forms
-------------------------------------
Loads form templates from /Forms (outside backend)
Generates filled Excel files into /filled_forms (outside backend)
"""

import datetime
import shutil
import tempfile
from pathlib import Path
from typing import Optional
import subprocess

from pydantic import BaseModel, Field
from enum import Enum
from openpyxl import load_workbook


# ====================================================================
# PATH RESOLUTION — GLOBAL PROJECT FOLDERS
# ====================================================================

BACKEND_DIR = Path(__file__).resolve().parents[1]    # backend/
PROJECT_ROOT = BACKEND_DIR.parent                   # project root

FORMS_DIR = PROJECT_ROOT / "Forms"                  # templates folder
FILLED_FORMS_DIR = PROJECT_ROOT / "filled_forms"    # output folder

FILLED_FORMS_DIR.mkdir(exist_ok=True)



# ====================================================================
# ENUMS & SCHEMAS
# ====================================================================

class LeaveType(str, Enum):
    ANNUAL = "annual"
    ACCIDENTAL = "accidental"
    MARRIAGE = "marriage"
    OTHER = "other"


class StaffCategory(str, Enum):
    ACADEMIC = "academic"
    NON_ACADEMIC = "non_academic"


class EmploymentType(str, Enum):
    FULL_TIME = "full_time"
    PART_TIME = "part_time"


class SalaryPaymentType(str, Enum):
    NORMAL = "normal"
    UNPAID = "unpaid"


class ApprovalStatus(str, Enum):
    YES = "yes"
    NO = "no"


class EmployeeName(BaseModel):
    first_name: str
    middle_name: Optional[str] = None
    last_name: str


class CompensationDetails(BaseModel):
    for_day: Optional[str] = None
    reason: Optional[str] = None


class SignatureEntry(BaseModel):
    name: Optional[str] = None
    date: Optional[datetime.date] = None


class HRApproval(BaseModel):
    approved: Optional[ApprovalStatus] = None
    reason: Optional[str] = None


class HRDetails(BaseModel):
    current_balance_days: Optional[int] = None
    leave_applied_days: Optional[int] = None
    salary_payment: Optional[SalaryPaymentType] = None
    extra_notes: Optional[str] = None
    hr_approval: Optional[HRApproval] = None
    final_approval: Optional[HRApproval] = None


class AnnualLeaveRequest(BaseModel):
    leave_type: LeaveType
    other_leave_description: Optional[str] = None

    compensation: Optional[CompensationDetails] = None

    employee_name: EmployeeName
    employee_id: str

    staff_category: StaffCategory
    employment_type: EmploymentType

    faculty: Optional[str] = None
    department: str

    leave_from: datetime.date
    leave_to: datetime.date
    total_leave_days: int = Field(..., ge=1)

    employee_signature: Optional[SignatureEntry] = None
    replacement_person: Optional[SignatureEntry] = None
    head_of_department: Optional[SignatureEntry] = None
    faculty_dean: Optional[SignatureEntry] = None

    hr_details: Optional[HRDetails] = None



# ====================================================================
# TEMPLATE MAP — ALL FORM FILES IN /Forms
# ====================================================================

FORM_TEMPLATES = {
    "annual": FORMS_DIR / "Annual_Leave_Request.xlsx",
    "accidental": FORMS_DIR / "Annual_Leave_Request.xlsx",
    "marriage": FORMS_DIR / "Annual_Leave_Request.xlsx",
    "excuse": FORMS_DIR / "Excuse_form.xlsx",
    "maternity": FORMS_DIR / "Maternity_form.xlsx",
    "mission": FORMS_DIR / "Mission_form.xlsx",
    "attendance": FORMS_DIR / "Incomplete_form.xlsx",
}


# ====================================================================
# HELPERS
# ====================================================================

def format_date(d: Optional[datetime.date]) -> str:
    return "" if d is None else d.strftime("%d/%m/%Y")


def convert_xls_to_xlsx(xls_path: str) -> str:
    """LibreOffice conversion for old .xls templates."""
    xls_path = Path(xls_path)
    with tempfile.TemporaryDirectory() as tmpdir:
        result = subprocess.run([
            "libreoffice", "--headless", "--convert-to", "xlsx",
            "--outdir", tmpdir, str(xls_path)
        ], capture_output=True, text=True)

        if result.returncode != 0:
            raise RuntimeError(f"Conversion failed: {result.stderr}")

        converted = Path(tmpdir) / f"{xls_path.stem}.xlsx"
        output_file = tempfile.gettempdir() + f"/{xls_path.stem}_converted.xlsx"
        shutil.copy(converted, output_file)
        return output_file



# ====================================================================
# MAIN PUBLIC FUNCTION
# ====================================================================

def fill_excel_form(form_type: str, parsed_data: dict, user_id: int) -> dict:
    """
    Main function used by the backend tool to generate filled forms.

    Returns { "file_path": "/abs/path/to/generated.xlsx" }
    """

    # Ensure template exists
    if form_type not in FORM_TEMPLATES:
        return {"error": "invalid_form_type"}

    template_path = FORM_TEMPLATES[form_type]

    if not template_path.exists():
        return {"error": "template_not_found", "detail": str(template_path)}

    # Validate data with schema
    try:
        data = AnnualLeaveRequest(**parsed_data)
    except Exception as e:
        return {"error": "validation_error", "detail": str(e)}

    # Output filename
    output_file = FILLED_FORMS_DIR / f"user_{user_id}_{form_type}_form.xlsx"

    # Convert template if needed
    if template_path.suffix == ".xls":
        template_path = Path(convert_xls_to_xlsx(str(template_path)))

    # Copy template → output
    shutil.copy(template_path, output_file)

    wb = load_workbook(output_file)
    ws = wb.active

    def write(r, c, value):
        ws.cell(row=r, column=c).value = value

    # Fill employee info
    full_name = f"{data.employee_name.first_name} " \
                f"{(data.employee_name.middle_name + ' ') if data.employee_name.middle_name else ''}" \
                f"{data.employee_name.last_name}"

    write(9, 3, full_name)
    write(14, 2, data.employee_id)

    # Leave type selection
    type_cells = {
        "annual": (5, 1, "☑ Annual"),
        "accidental": (5, 4, "☑ Accidental"),
        "marriage": (5, 9, "☑ Marriage"),
        "other": (5, 14, f"☑ Other: {data.other_leave_description or ''}")
    }

    r, c, text = type_cells[data.leave_type.value]
    write(r, c, text)

    # Staff category
    if data.staff_category == StaffCategory.ACADEMIC:
        write(17, 1 if data.employment_type == EmploymentType.FULL_TIME else 5,
              f"☑ {'Full time' if data.employment_type == EmploymentType.FULL_TIME else 'Part time'}")
        write(18, 1, f"Faculty: {data.faculty or ''}\nDepartment: {data.department}")
    else:
        write(17, 10 if data.employment_type == EmploymentType.FULL_TIME else 15,
              f"☑ {'Full time' if data.employment_type == EmploymentType.FULL_TIME else 'Part time'}")
        write(18, 10, f"Department: {data.department}")

    # Leave period
    write(19, 1, f"Leave from: {format_date(data.leave_from)}")
    write(19, 10, f"to: {format_date(data.leave_to)}")
    write(20, 4, str(data.total_leave_days))

    # Save final file
    wb.save(output_file)

    return {"file_path": str(output_file)}
